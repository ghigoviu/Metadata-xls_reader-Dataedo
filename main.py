from Utils.HandleCSV import HandleCSV
from openpyxl.reader.excel import load_workbook

def append_tables(act_list, new_table):
    for table_row in new_table[1:]:
        act_list.append(table_row)

def get_titles(num_worksheet, title_row = 11):
    row_ini = title_row
    column_ini = 1
    ws = wb.worksheets[num_worksheet]
    column_act = column_ini
    cl = ws.cell(row=row_ini, column=column_ini)
    while cl.value is not None:
        column_act += 1
        cl = ws.cell(row=row_ini, column=column_act)

    titles = ["Type", "BD", "Name"]
    num_columns = column_act
    for row in wb.worksheets[num_worksheet].iter_rows(min_row=title_row, max_col=num_columns - 1, max_row=title_row):
        for cell in row:
            titles.append(cell.value)
    return titles

def append_relationships(ws_act):
    flag = False
    new_list = []
    row_ini = 13
    relationships_start = row_ini
    for row in ws_act.iter_rows(min_row=row_ini, max_row=100, min_col=0, max_col=1):
        for cell in row:
            if cell.value and "Relationships" in str(cell.value):
                flag = True
                relationships_start = row_ini
                break
            row_ini += 1
    if flag:
        for row in ws_act.iter_rows(min_row=relationships_start+1, max_row=500, min_col=1, max_col=11):
            new_row = ["Relationship", ws_act.cell(row=3, column=2).value, ws_act.cell(row=5, column=2).value]
            for cell in row:
                new_row.append(cell.value)
            if row[0].value is None:
                break
            new_list.append(new_row)
        return new_list
    else:
        return flag

if __name__ == '__main__':
    wb = load_workbook('resources/Dataedo Data Dictionary.xlsx')

    tables_list = [get_titles(0)]
    functions_list = []
    procedures_list = []
    views_list = []
    relationships_list = []
    relationships_list.append(["Type", "BD", "Name", "Foreign database", "Foreign table", "Type", "Primary database",
                          "Primary table", "Join", "Title", "Relationship name", "Description", "Owner"])
    all_list = []
    # for sheet in wb.worksheets[4:]:
    for sheet in wb.worksheets:
        ws = sheet
        row_ini = 11
        column_ini = 1
        # Cuántas columnas tiene la tabla?
        row_act = row_ini
        column_act = column_ini
        cl = ws.cell(row=row_ini, column=column_ini)
        while cl.value is not None:
            column_act += 1
            cl = ws.cell(row=row_ini, column=column_act)

        # Cuántas filas tiene la tabla?
        cl = ws.cell(row=row_ini, column=column_ini)
        while cl.value is not None:
            row_act += 1
            cl = ws.cell(row=row_act, column=column_ini)

        # Obtener el dato de toda la tabla
        num_filas = row_act
        num_columns = column_act
        new_list = []
        new_row = ["Type", "Documentation", "Name"]
        titles_row = 11
        for row in ws.iter_rows(min_row=titles_row, max_col=num_columns-1, max_row=titles_row):
            for cell in row:
                new_row.append(cell.value)
        new_list.append(new_row)

        for row in ws.iter_rows(min_row=titles_row + 1, max_col=num_columns-1, max_row=num_filas-1):
            new_row = [ws.cell(row=1, column=1).value, ws.cell(row=3, column=2).value, ws.cell(row=5, column=2).value]
            for cell in row:
                new_row.append(cell.value)
            new_list.append(new_row)

        if ws.cell(row=1, column=1).value == "Table:":
            append_tables(tables_list, new_list)
            rel_list = append_relationships(ws)
            if rel_list is not False:
                append_tables(relationships_list, rel_list)
        elif ws.cell(row=1, column=1).value == "View:":
            append_tables(views_list, new_list)
        elif ws.cell(row=1, column=1).value == "Function:":
            append_tables(functions_list, new_list)
        elif ws.cell(row=1, column=1).value == "Procedure:":
            append_tables(procedures_list, new_list)
        for e in new_list:
            all_list.append(e)

    HandleCSV.write_csv("./resources/all_list_full.csv", all_list)
    HandleCSV.write_csv("./resources/table_list_full.csv", tables_list)
    HandleCSV.write_csv("./resources/views_list_full.csv", views_list)
    HandleCSV.write_csv("./resources/functions_list_full.csv", functions_list)
    HandleCSV.write_csv("./resources/procedures_list_full.csv", procedures_list)
    HandleCSV.write_csv("./resources/relationships_list_full.csv", relationships_list)
